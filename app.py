"""
NFC Attendance System - Backend
Excel-based database + pywebview standalone window
Supports multiple services (خدمات) and stages (مراحل) under Sunday School.
Master admin = the desktop (pywebview) session.
"""

from flask import Flask, jsonify, request, send_from_directory, session # pyre-ignore
from werkzeug.security import generate_password_hash, check_password_hash # pyre-ignore
from functools import wraps
from flask_cors import CORS # pyre-ignore
from openpyxl import Workbook, load_workbook # pyre-ignore
import threading
import os
import sys
import re
import random
import base64
from urllib import parse, request as urlrequest, error as urlerror
from datetime import datetime, date, timedelta

app = Flask(__name__)
CORS(app)
app.permanent_session_lifetime = timedelta(days=365)
app.secret_key = os.environ.get('SECRET_KEY', 'nfc-attendance-secret-key-change-in-production')

# When running as a PyInstaller .exe, use the exe's directory for runtime files
# and the bundled _MEIPASS directory for static assets
if getattr(sys, 'frozen', False):
    RUN_DIR = os.path.dirname(sys.executable)
    BASE_DIR = getattr(sys, '_MEIPASS', RUN_DIR)
else:
    RUN_DIR = os.path.dirname(os.path.abspath(__file__))
    BASE_DIR = RUN_DIR

EXCEL_PATH = os.path.join(RUN_DIR, 'attendance_data.xlsx')
_excel_lock = threading.Lock()
_otp_lock = threading.Lock()
OTP_STORE = {}

OTP_TTL_MINUTES = int(os.environ.get('OTP_TTL_MINUTES', '5'))
OTP_MAX_ATTEMPTS = int(os.environ.get('OTP_MAX_ATTEMPTS', '5'))

# Master admin secret — sent as a special header from pywebview
MASTER_ADMIN_SECRET = os.environ.get('MASTER_ADMIN_SECRET', 'pywebview-master-admin-2024')

# Master admin phone number — this supervisor gets admin privileges on login
MASTER_ADMIN_PHONE = '01274767185'
MASTER_ADMIN_PASSWORD = '1234'


# ===== UTILITY FUNCTIONS (must be defined before _init_excel) =====

def normalize_phone(phone):
    """Keep only digits to normalize phone numbers before comparison/storage."""
    return ''.join(ch for ch in str(phone) if ch.isdigit())


def is_valid_phone(phone):
    """Validate Egyptian mobile numbers (11 digits starting with 01)."""
    return bool(re.fullmatch(r"01\d{9}", phone))


def to_e164_eg(phone):
    """Convert local Egyptian mobile number to E.164 format for SMS providers."""
    normalized = normalize_phone(phone)
    if normalized.startswith('01') and len(normalized) == 11:
        return '+2' + normalized
    return normalized


# ===== EXCEL DATABASE LAYER =====

def _init_excel():
    """Create the Excel file with proper sheets and headers if it doesn't exist."""
    if os.path.exists(EXCEL_PATH):
        # Migrate: ensure new sheets exist in older files
        _migrate_excel()
        return
    wb = Workbook()
    # Services sheet
    ws_svc = wb.active
    ws_svc.title = 'Services'
    ws_svc.append(['id', 'name', 'created_at'])
    # Seed default service: مدارس الاحد
    ws_svc.append([1, 'مدارس الاحد', datetime.now().isoformat()])

    # Stages sheet
    ws_stg = wb.create_sheet('Stages')
    ws_stg.append(['id', 'service_id', 'name', 'created_at'])
    # Seed default stages for مدارس الاحد (service_id=1)
    default_stages = ['ملائكة', 'اولي وتانية', 'رابعة وخامسة وسادسة', 'اعدادي', 'ثانوي']
    for i, stage_name in enumerate(default_stages, start=1):
        ws_stg.append([i, 1, stage_name, datetime.now().isoformat()])

    # Supervisors sheet (now with service_id and stage_id)
    ws_sup = wb.create_sheet('Supervisors')
    ws_sup.append(['id', 'name', 'phone', 'password_hash', 'service_id', 'stage_id', 'created_at'])
    # Seed master admin supervisor
    ws_sup.append([1, 'المدير العام', MASTER_ADMIN_PHONE,
                   generate_password_hash(MASTER_ADMIN_PASSWORD), 0, 0,
                   datetime.now().isoformat()])

    # Employees sheet (now with service_id and stage_id)
    ws_emp = wb.create_sheet('Employees')
    ws_emp.append(['id', 'nfc_uid', 'name', 'birthdate', 'address', 'phone',
                   'email', 'department', 'parent_phone', 'confession_father',
                   'photo_url', 'service_id', 'stage_id', 'created_at'])

    # Attendance sheet
    ws_att = wb.create_sheet('Attendance')
    ws_att.append(['id', 'employee_id', 'supervisor_id', 'scan_time', 'status', 'notes'])

    # Visits (افتقادات) sheet
    ws_vis = wb.create_sheet('Visits')
    ws_vis.append(['id', 'employee_id', 'supervisor_id', 'visit_date', 'notes', 'created_at'])

    wb.save(EXCEL_PATH)


def _migrate_excel():
    """Add missing sheets/columns to an existing Excel file for backward compatibility."""
    with _excel_lock:
        wb = _load_wb()
        changed = False

        # Ensure Services sheet exists
        if 'Services' not in wb.sheetnames:
            ws_svc = wb.create_sheet('Services')
            ws_svc.append(['id', 'name', 'created_at'])
            ws_svc.append([1, 'مدارس الاحد', datetime.now().isoformat()])
            changed = True

        # Ensure Stages sheet exists
        if 'Stages' not in wb.sheetnames:
            ws_stg = wb.create_sheet('Stages')
            ws_stg.append(['id', 'service_id', 'name', 'created_at'])
            default_stages = ['ملائكة', 'اولي وتانية', 'رابعة وخامسة وسادسة', 'اعدادي', 'ثانوي']
            for i, stage_name in enumerate(default_stages, start=1):
                ws_stg.append([i, 1, stage_name, datetime.now().isoformat()])
            changed = True

        # Ensure Visits sheet exists
        if 'Visits' not in wb.sheetnames:
            ws_vis = wb.create_sheet('Visits')
            ws_vis.append(['id', 'employee_id', 'supervisor_id', 'visit_date', 'notes', 'created_at'])
            changed = True

        # Add service_id, stage_id columns to Supervisors if missing
        if 'Supervisors' in wb.sheetnames:
            ws_sup = wb['Supervisors']
            headers = _get_headers(ws_sup)
            if 'service_id' not in headers:
                col = len(headers) + 1
                ws_sup.cell(row=1, column=col, value='service_id')
                changed = True
            if 'stage_id' not in headers:
                col = len(_get_headers(ws_sup)) + 1
                ws_sup.cell(row=1, column=col, value='stage_id')
                changed = True

        # Add service_id, stage_id columns to Employees if missing
        if 'Employees' in wb.sheetnames:
            ws_emp = wb['Employees']
            headers = _get_headers(ws_emp)
            if 'service_id' not in headers:
                col = len(headers) + 1
                ws_emp.cell(row=1, column=col, value='service_id')
                changed = True
            if 'stage_id' not in headers:
                col = len(_get_headers(ws_emp)) + 1
                ws_emp.cell(row=1, column=col, value='stage_id')
                changed = True

        # Ensure master admin supervisor exists
        if 'Supervisors' in wb.sheetnames:
            ws_sup = wb['Supervisors']
            sup_headers = _get_headers(ws_sup)
            phone_col = sup_headers.index('phone') + 1
            admin_exists = False
            for row_num in range(2, ws_sup.max_row + 1):
                val = ws_sup.cell(row=row_num, column=phone_col).value
                if val is not None and normalize_phone(str(val)) == MASTER_ADMIN_PHONE:
                    admin_exists = True
                    break
            if not admin_exists:
                new_id = _next_id(ws_sup)
                ws_sup.append([new_id, 'المدير العام', MASTER_ADMIN_PHONE,
                               generate_password_hash(MASTER_ADMIN_PASSWORD),
                               0, 0, datetime.now().isoformat()])
                changed = True

        if changed:
            wb.save(EXCEL_PATH)
        wb.close()


def _load_wb():
    return load_workbook(EXCEL_PATH)


def _next_id(ws):
    """Get next auto-increment ID for a sheet."""
    max_id = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0] is not None:
            try:
                max_id = max(max_id, int(row[0]))
            except (ValueError, TypeError):
                pass
    return max_id + 1


def _sheet_to_dicts(ws):
    """Convert a worksheet to a list of dicts using the header row as keys."""
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 1:
        return []
    headers = [str(h) for h in rows[0]]
    result = []
    for i in range(1, len(rows)):
        row = rows[i]
        if row[0] is None:
            continue
        d = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            d[h] = val if val is not None else ''
        result.append(d)
    return result


def _find_row_by_id(ws, record_id):
    """Find row number (1-indexed) by ID in column 1. Returns None if not found."""
    for row_num in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_num, column=1).value
        if cell_val is not None and int(cell_val) == int(record_id):
            return row_num
    return None


def _get_headers(ws):
    return [str(ws.cell(row=1, column=c).value) for c in range(1, ws.max_column + 1)]


def _row_to_dict(ws, row_num):
    headers = _get_headers(ws)
    d = {}
    for i, h in enumerate(headers):
        val = ws.cell(row=row_num, column=i + 1).value
        d[h] = val if val is not None else ''
    return d


def _safe_int(val, default=0):
    """Safely convert a value to int."""
    if val is None or val == '':
        return default
    try:
        return int(val)
    except (ValueError, TypeError):
        return default


# ===== INIT =====
_init_excel()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'supervisor_id' not in session and not _is_master_admin():
            return jsonify({'error': 'Authentication required'}), 401
        return f(*args, **kwargs)
    return decorated


def master_admin_required(f):
    """Only the desktop (pywebview) master admin can call these endpoints."""
    @wraps(f)
    def decorated(*args, **kwargs):
        if not _is_master_admin():
            return jsonify({'error': 'Master admin access required'}), 403
        return f(*args, **kwargs)
    return decorated


def _is_master_admin():
    """Check if the current request is from the master admin (pywebview desktop)."""
    # Method 1: special header
    secret = request.headers.get('X-Master-Admin-Secret', '')
    if secret == MASTER_ADMIN_SECRET:
        return True
    # Method 2: session flag set by /api/admin/login
    return session.get('is_master_admin', False)


def get_nearest_friday(d=None):
    if d is None:
        d = date.today()
    days_since_friday = (d.weekday() - 4) % 7
    return d - timedelta(days=days_since_friday)







def generate_otp_code():
    return f"{random.randint(0, 999999):06d}"


def cleanup_expired_otps():
    now = datetime.now()
    with _otp_lock:
        expired_phones = [p for p, data in OTP_STORE.items() if data['expires_at'] < now]
        for p in expired_phones:
            OTP_STORE.pop(p, None)


def send_sms_otp(phone, otp_code):
    """
    Send OTP via SMS.
    Modes:
    - SMS_MODE=twilio with TWILIO_* vars for real SMS
    - default simulated mode (no external dependency)
    """
    sms_mode = os.environ.get('SMS_MODE', 'simulated').strip().lower()
    if sms_mode != 'twilio':
        return {'sent': True, 'simulated': True, 'message': 'SIMULATED_SMS'}

    account_sid = os.environ.get('TWILIO_ACCOUNT_SID', '').strip()
    auth_token = os.environ.get('TWILIO_AUTH_TOKEN', '').strip()
    from_number = os.environ.get('TWILIO_FROM_NUMBER', '').strip()
    to_number = to_e164_eg(phone)
    if not account_sid or not auth_token or not from_number:
        return {'sent': False, 'simulated': False, 'error': 'Twilio credentials are not configured'}

    endpoint = f"https://api.twilio.com/2010-04-01/Accounts/{account_sid}/Messages.json"
    payload = parse.urlencode({
        'From': from_number,
        'To': to_number,
        'Body': f'Your OTP code is: {otp_code}. It expires in {OTP_TTL_MINUTES} minutes.'
    }).encode('utf-8')

    basic_auth = base64.b64encode(f"{account_sid}:{auth_token}".encode('utf-8')).decode('utf-8')
    req = urlrequest.Request(endpoint, data=payload, method='POST')
    req.add_header('Authorization', f'Basic {basic_auth}')
    req.add_header('Content-Type', 'application/x-www-form-urlencoded')

    try:
        with urlrequest.urlopen(req, timeout=15) as resp:
            if 200 <= resp.status < 300:
                return {'sent': True, 'simulated': False}
            return {'sent': False, 'simulated': False, 'error': f'SMS provider returned {resp.status}'}
    except urlerror.HTTPError as exc:
        return {'sent': False, 'simulated': False, 'error': f'SMS provider HTTP error {exc.code}'}
    except Exception as exc:
        return {'sent': False, 'simulated': False, 'error': str(exc)}


def _get_next_birthday(birthdate_str):
    """Given a birthdate string (YYYY-MM-DD), return the next birthday as ISO string."""
    if not birthdate_str or birthdate_str == '':
        return ''
    try:
        bd_str = str(birthdate_str).strip().split('T')[0].split(' ')[0]
        bd = datetime.strptime(bd_str, '%Y-%m-%d').date()
    except (ValueError, TypeError):
        return ''
    today = date.today()
    this_year_bd = bd.replace(year=today.year)
    if this_year_bd < today:
        this_year_bd = bd.replace(year=today.year + 1)
    return this_year_bd.isoformat()


def _get_last_visit_date(employee_id, visits):
    """Get the most recent visit date for an employee from visits list."""
    emp_visits = [v for v in visits if _safe_int(v.get('employee_id')) == employee_id]
    if not emp_visits:
        return ''
    emp_visits.sort(key=lambda v: str(v.get('visit_date', '')), reverse=True)
    return str(emp_visits[0].get('visit_date', ''))


# ===== STATIC FILE SERVING =====

@app.route('/')
def serve_index():
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/css/<path:path>')
def serve_css(path):
    return send_from_directory(os.path.join(BASE_DIR, 'css'), path)

@app.route('/js/<path:path>')
def serve_js(path):
    return send_from_directory(os.path.join(BASE_DIR, 'js'), path)

@app.route('/manifest.json')
def serve_manifest():
    return send_from_directory(BASE_DIR, 'manifest.json')

@app.route('/sw.js')
def serve_sw():
    return send_from_directory(BASE_DIR, 'sw.js')


# ===== MASTER ADMIN ENDPOINTS =====

@app.route('/api/admin/login', methods=['POST'])
def admin_login():
    """Login as master admin using a secret key (from pywebview desktop)."""
    data = request.json or {}
    secret = data.get('secret', '')
    if secret == MASTER_ADMIN_SECRET:
        session['is_master_admin'] = True
        session['supervisor_id'] = -1  # special admin ID
        return jsonify({'message': 'OK', 'is_master_admin': True})
    return jsonify({'error': 'Invalid admin secret'}), 401


@app.route('/api/admin/check', methods=['GET'])
@app.route('/api/ping', methods=['GET'])
def ping():
    return jsonify({'nfc_server': True, 'message': 'pong'})

def admin_check():
    """Check if current session is master admin."""
    return jsonify({'is_master_admin': _is_master_admin()})


# ===== SERVICES ENDPOINTS (master admin only for create/delete) =====

@app.route('/api/services', methods=['GET'])
def list_services():
    """List all services. No auth required so registration page can show them."""
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Services']
        services = _sheet_to_dicts(ws)
        wb.close()
    result = []
    for s in services:
        result.append({
            'id': _safe_int(s['id']),
            'name': s['name'],
            'created_at': s.get('created_at', ''),
        })
    result.sort(key=lambda x: x['name'])
    return jsonify(result)


@app.route('/api/services', methods=['POST'])
@master_admin_required
def create_service():
    data = request.json
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'error': 'Service name required'}), 400
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Services']
        # Check uniqueness
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and str(row[1]).strip() == name:
                wb.close()
                return jsonify({'error': 'Service already exists'}), 409
        new_id = _next_id(ws)
        ws.append([new_id, name, datetime.now().isoformat()])
        wb.save(EXCEL_PATH)
        wb.close()
    return jsonify({'message': 'OK', 'service': {'id': new_id, 'name': name}}), 201


@app.route('/api/services/<int:svc_id>', methods=['DELETE'])
@master_admin_required
def delete_service(svc_id):
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Services']
        row = _find_row_by_id(ws, svc_id)
        if row:
            ws.delete_rows(row)
        # Also delete all stages of this service
        ws_stg = wb['Stages']
        rows_to_delete = []
        headers = _get_headers(ws_stg)
        svc_col = headers.index('service_id') + 1
        for r in range(2, ws_stg.max_row + 1):
            val = ws_stg.cell(row=r, column=svc_col).value
            if val is not None and _safe_int(val) == svc_id:
                rows_to_delete.append(r)
        for r in reversed(rows_to_delete):
            ws_stg.delete_rows(r)
        wb.save(EXCEL_PATH)
        wb.close()
    return jsonify({'message': 'OK'})


# ===== STAGES ENDPOINTS (master admin only for create/delete) =====

@app.route('/api/stages', methods=['GET'])
def list_stages():
    """List all stages, optionally filtered by service_id. No auth required for registration."""
    service_id = request.args.get('service_id')
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Stages']
        stages = _sheet_to_dicts(ws)
        wb.close()
    result = []
    for s in stages:
        if service_id and _safe_int(s.get('service_id')) != _safe_int(service_id):
            continue
        result.append({
            'id': _safe_int(s['id']),
            'service_id': _safe_int(s.get('service_id')),
            'name': s['name'],
            'created_at': s.get('created_at', ''),
        })
    result.sort(key=lambda x: x['name'])
    return jsonify(result)


@app.route('/api/stages', methods=['POST'])
@master_admin_required
def create_stage():
    data = request.json
    name = data.get('name', '').strip()
    service_id = _safe_int(data.get('service_id'))
    if not name or not service_id:
        return jsonify({'error': 'Stage name and service_id required'}), 400
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Stages']
        # Check uniqueness within service
        for row in ws.iter_rows(min_row=2, values_only=True):
            if (row[0] is not None and _safe_int(row[1]) == service_id
                    and str(row[2]).strip() == name):
                wb.close()
                return jsonify({'error': 'Stage already exists in this service'}), 409
        new_id = _next_id(ws)
        ws.append([new_id, service_id, name, datetime.now().isoformat()])
        wb.save(EXCEL_PATH)
        wb.close()
    return jsonify({'message': 'OK', 'stage': {'id': new_id, 'service_id': service_id, 'name': name}}), 201


@app.route('/api/stages/<int:stage_id>', methods=['DELETE'])
@master_admin_required
def delete_stage(stage_id):
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Stages']
        row = _find_row_by_id(ws, stage_id)
        if row:
            ws.delete_rows(row)
        wb.save(EXCEL_PATH)
        wb.close()
    return jsonify({'message': 'OK'})


# ===== AUTH ENDPOINTS =====

@app.route('/api/auth/register', methods=['POST'])
def register():
    data = request.json
    name = data.get('name', '').strip()
    phone = normalize_phone(data.get('phone', '').strip())
    password = data.get('password', '')
    service_id = _safe_int(data.get('service_id'))
    stage_id = _safe_int(data.get('stage_id'))
    if not name or not phone or not password:
        return jsonify({'error': 'All fields required'}), 400
    if not is_valid_phone(phone):
        return jsonify({'error': 'Invalid phone number. Use 11 digits starting with 01'}), 400
    if len(password) < 4:
        return jsonify({'error': 'Password must be at least 4 characters'}), 400
    if not service_id or not stage_id:
        return jsonify({'error': 'Service and stage are required'}), 400
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Supervisors']
        # Check uniqueness
        headers = _get_headers(ws)
        phone_col = headers.index('phone') + 1
        for row_num in range(2, ws.max_row + 1):
            val = ws.cell(row=row_num, column=phone_col).value
            if val is not None and normalize_phone(val) == phone:
                wb.close()
                return jsonify({'error': 'Phone already registered'}), 409
        new_id = _next_id(ws)
        ws.append([new_id, name, phone, generate_password_hash(password),
                   service_id, stage_id, datetime.now().isoformat()])
        wb.save(EXCEL_PATH)
        wb.close()
    session.permanent = True
    session['supervisor_id'] = new_id
    return jsonify({'message': 'OK', 'supervisor': {
        'id': new_id, 'name': name, 'phone': phone,
        'service_id': service_id, 'stage_id': stage_id
    }}), 201


@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.json
    phone = normalize_phone(data.get('phone', '').strip())
    password = data.get('password', '')
    if not is_valid_phone(phone):
        return jsonify({'error': 'Invalid phone number'}), 400
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Supervisors']
        supervisors = _sheet_to_dicts(ws)
        wb.close()
    supervisor = {}
    for s in supervisors:
        if normalize_phone(s.get('phone', '')) == phone:
            supervisor = s
            break
    if not supervisor or not check_password_hash(supervisor.get('password_hash', ''), password):
        return jsonify({'error': 'Invalid credentials'}), 401
    session.permanent = True
    session['supervisor_id'] = int(supervisor['id'])
    # Grant master admin if this is the admin phone
    is_admin = (phone == MASTER_ADMIN_PHONE)
    if is_admin:
        session['is_master_admin'] = True
    return jsonify({
        'message': 'OK',
        'is_master_admin': is_admin,
        'supervisor': {
            'id': int(supervisor['id']), 'name': supervisor['name'],
            'phone': supervisor['phone'],
            'service_id': _safe_int(supervisor.get('service_id')),
            'stage_id': _safe_int(supervisor.get('stage_id')),
        }
    })


@app.route('/api/auth/forgot-password/request-otp', methods=['POST'])
def forgot_password_request_otp():
    data = request.json
    phone = normalize_phone(data.get('phone', '').strip())
    if not phone:
        return jsonify({'error': 'Phone is required'}), 400
    if not is_valid_phone(phone):
        return jsonify({'error': 'Invalid phone number. Use 11 digits starting with 01'}), 400

    cleanup_expired_otps()

    # Check phone exists
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Supervisors']
        phone_exists = False
        headers = _get_headers(ws)
        phone_col = headers.index('phone') + 1
        for row_num in range(2, ws.max_row + 1):
            row_phone = normalize_phone(ws.cell(row=row_num, column=phone_col).value)
            if row_phone == phone:
                phone_exists = True
                break
        wb.close()

    if not phone_exists:
        return jsonify({'error': 'Phone number not found'}), 404

    otp_code = generate_otp_code()
    expires_at = datetime.now() + timedelta(minutes=OTP_TTL_MINUTES)
    with _otp_lock:
        OTP_STORE[phone] = {
            'otp_code': otp_code,
            'expires_at': expires_at,
            'attempts': 0
        }

    sms_result = send_sms_otp(phone, otp_code)
    if not sms_result.get('sent'):
        return jsonify({'error': sms_result.get('error', 'Failed to send OTP')}), 500

    response: dict = {
        'message': f'OTP has been sent. It expires in {OTP_TTL_MINUTES} minutes.'
    }
    if sms_result.get('simulated'):
        response['simulated'] = True
        response['otp_preview'] = otp_code

    return jsonify(response)


@app.route('/api/auth/forgot-password/verify-otp', methods=['POST'])
def forgot_password_verify_otp():
    data = request.json
    phone = normalize_phone(data.get('phone', '').strip())
    otp_code = str(data.get('otp_code', '')).strip()
    new_password = data.get('new_password', '')

    if not phone or not otp_code or not new_password:
        return jsonify({'error': 'Phone, OTP, and new password are required'}), 400
    if not is_valid_phone(phone):
        return jsonify({'error': 'Invalid phone number. Use 11 digits starting with 01'}), 400
    if not re.fullmatch(r'\d{6}', otp_code):
        return jsonify({'error': 'OTP must be 6 digits'}), 400
    if len(new_password) < 4:
        return jsonify({'error': 'Password must be at least 4 characters'}), 400

    cleanup_expired_otps()

    with _otp_lock:
        otp_data = OTP_STORE.get(phone)
        if not otp_data:
            return jsonify({'error': 'OTP not found or expired. Request a new code.'}), 400
        if otp_data['attempts'] >= OTP_MAX_ATTEMPTS:
            OTP_STORE.pop(phone, None)
            return jsonify({'error': 'Too many attempts. Request a new OTP.'}), 429
        if otp_data['otp_code'] != otp_code:
            otp_data['attempts'] += 1
            return jsonify({'error': 'Invalid OTP code'}), 400

    with _excel_lock:
        wb = _load_wb()
        ws = wb['Supervisors']
        headers = _get_headers(ws)
        phone_col = headers.index('phone') + 1
        pw_col = headers.index('password_hash') + 1
        target_row = None
        for row_num in range(2, ws.max_row + 1):
            row_phone = normalize_phone(ws.cell(row=row_num, column=phone_col).value)
            if row_phone == phone:
                target_row = row_num
                break
        if not target_row:
            wb.close()
            with _otp_lock:
                OTP_STORE.pop(phone, None)
            return jsonify({'error': 'Phone number not found'}), 404

        ws.cell(row=target_row, column=pw_col, value=generate_password_hash(new_password))
        wb.save(EXCEL_PATH)
        wb.close()

    with _otp_lock:
        OTP_STORE.pop(phone, None)

    return jsonify({'message': 'Password has been reset successfully'})


@app.route('/api/auth/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'message': 'OK'})


@app.route('/api/auth/me', methods=['GET'])
def auth_me():
    if _is_master_admin():
        return jsonify({
            'authenticated': True,
            'is_master_admin': True,
            'supervisor': {
                'id': -1, 'name': 'المدير العام', 'phone': '',
                'service_id': 0, 'stage_id': 0,
            }
        })
    if 'supervisor_id' not in session:
        return jsonify({'authenticated': False}), 401
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Supervisors']
        supervisors = _sheet_to_dicts(ws)
        wb.close()
    supervisor = {}
    for s in supervisors:
        if int(s.get('id', 0)) == session['supervisor_id']:
            supervisor = s
            break
    if not supervisor:
        session.clear()
        return jsonify({'authenticated': False}), 401
    return jsonify({'authenticated': True, 'supervisor': {
        'id': int(supervisor.get('id', 0)), 'name': supervisor.get('name', ''),
        'phone': supervisor.get('phone', ''),
        'service_id': _safe_int(supervisor.get('service_id')),
        'stage_id': _safe_int(supervisor.get('stage_id')),
    }})


@app.route('/api/supervisors', methods=['GET'])
@login_required
def list_supervisors():
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Supervisors']
        supervisors = _sheet_to_dicts(ws)
        wb.close()
    result = sorted([{
        'id': int(s['id']), 'name': s['name'], 'phone': s['phone'],
        'service_id': _safe_int(s.get('service_id')),
        'stage_id': _safe_int(s.get('stage_id')),
        'created_at': s.get('created_at', '')
    } for s in supervisors], key=lambda x: x['name'])
    return jsonify(result)


@app.route('/api/supervisors/<int:sup_id>', methods=['GET'])
@login_required
def get_supervisor(sup_id):
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Supervisors']
        row = _find_row_by_id(ws, sup_id)
        if not row:
            wb.close()
            return jsonify({'error': 'Not found'}), 404
        sup = _row_to_dict(ws, row)
        
        # Count stats for this supervisor
        ws_att = wb['Attendance']
        attendance = _sheet_to_dicts(ws_att)
        wb.close()
        
    sup_dict = {
        'id': int(sup['id']), 'name': sup['name'], 'phone': sup['phone'],
        'service_id': _safe_int(sup.get('service_id')),
        'stage_id': _safe_int(sup.get('stage_id')),
        'created_at': sup.get('created_at', '')
    }
    
    total_scans = sum(1 for a in attendance if _safe_int(a.get('supervisor_id')) == sup_id)
    return jsonify({
        'supervisor': sup_dict,
        'stats': {'total_scans': total_scans}
    })

@app.route('/api/supervisors/<int:sup_id>', methods=['PUT'])
@master_admin_required
def update_supervisor(sup_id):
    data = request.json
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Supervisors']
        row_num = _find_row_by_id(ws, sup_id)
        if not row_num:
            wb.close()
            return jsonify({'error': 'Not found'}), 404
        
        headers = _get_headers(ws)
        old = _row_to_dict(ws, row_num)
        
        field_map = {
            'name': data.get('name', old['name']),
            'phone': normalize_phone(data.get('phone', old['phone'])),
            'service_id': _safe_int(data.get('service_id', old.get('service_id'))),
            'stage_id': _safe_int(data.get('stage_id', old.get('stage_id'))),
        }
        
        new_password = data.get('password', '').strip()
        if new_password:
            field_map['password_hash'] = generate_password_hash(new_password)
            
        for field, value in field_map.items():
            if field in headers:
                ws.cell(row=row_num, column=headers.index(field) + 1, value=value)
                
        wb.save(EXCEL_PATH)
        updated = _row_to_dict(ws, row_num)
        wb.close()
        
    return jsonify({'message': 'OK', 'supervisor': {
        'id': int(updated['id']), 'name': updated['name'], 'phone': updated['phone'],
        'service_id': _safe_int(updated.get('service_id')),
        'stage_id': _safe_int(updated.get('stage_id'))
    }})

@app.route('/api/supervisors/<int:sup_id>', methods=['DELETE'])
@master_admin_required
def delete_supervisor(sup_id):
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Supervisors']
        row = _find_row_by_id(ws, sup_id)
        if row:
            ws.delete_rows(row)
            
        # Optional: nullify or delete attendance/visits done by this supervisor
        # we'll skip for now to preserve history, or we could just set supervisor_id to null
        
        wb.save(EXCEL_PATH)
        wb.close()
    return jsonify({'message': 'OK'})

@app.route('/api/nfc/scan', methods=['POST'])
@login_required
def nfc_scan():
    data = request.json
    nfc_uid = data.get('nfc_uid', '').strip().upper()
    scan_date = data.get('date', date.today().isoformat())
    if not nfc_uid:
        return jsonify({'error': 'NFC UID required'}), 400
    with _excel_lock:
        wb = _load_wb()
        ws_emp = wb['Employees']
        employees = _sheet_to_dicts(ws_emp)
        employee = {}
        for e in employees:
            if str(e.get('nfc_uid', '')).upper() == nfc_uid:
                employee = e
                break
        if not employee:
            wb.close()
            return jsonify({'status': 'unknown', 'nfc_uid': nfc_uid, 'message': 'Unknown card'})
        # Check if already scanned today
        ws_att = wb['Attendance']
        attendance = _sheet_to_dicts(ws_att)
        day_start = scan_date + ' 00:00:00'
        day_end = scan_date + ' 23:59:59'
        existing = {}
        for a in attendance:
            scan_t = str(a.get('scan_time', ''))
            if (int(a.get('employee_id', 0)) == int(employee.get('id', 0)) and
                    day_start <= scan_t and scan_t <= day_end):
                existing = a
                break
        if existing:
            wb.close()
            emp_dict = {k: (int(v) if k == 'id' else v) for k, v in employee.items()}
            return jsonify({
                'status': 'already_scanned', 'employee': emp_dict,
                'scan_time': existing.get('scan_time', ''),
                'message': str(employee.get('name', '')) + ' already scanned'
            })
        # Record attendance
        now_time = datetime.now().strftime('%H:%M:%S')
        record_time = scan_date + ' ' + now_time
        new_id = _next_id(ws_att)
        sup_id = session.get('supervisor_id', -1)
        ws_att.append([new_id, int(employee['id']), sup_id, record_time, 'present', ''])
        wb.save(EXCEL_PATH)
        wb.close()
    emp_dict = {k: (int(v) if k == 'id' else v) for k, v in employee.items()}
    return jsonify({
        'status': 'recorded', 'employee': emp_dict,
        'scan_time': record_time, 'message': 'Recorded ' + employee['name']
    })


@app.route('/api/attendance/manual', methods=['POST'])
@login_required
def manual_attendance():
    data = request.json
    employee_id = data.get('employee_id')
    status = data.get('status', 'present')
    notes = data.get('notes', '')
    record_date = data.get('date', date.today().isoformat())
    if not employee_id:
        return jsonify({'error': 'Employee ID required'}), 400
    employee_id = int(employee_id)
    with _excel_lock:
        wb = _load_wb()
        ws_att = wb['Attendance']
        attendance = _sheet_to_dicts(ws_att)
        day_start = record_date + ' 00:00:00'
        day_end = record_date + ' 23:59:59'
        existing = {}
        for a in attendance:
            scan_t = str(a.get('scan_time', ''))
            if (int(a.get('employee_id', 0)) == employee_id and
                    day_start <= scan_t and scan_t <= day_end):
                existing = a
                break
        now_time = datetime.now().strftime('%H:%M:%S')
        record_time = record_date + ' ' + now_time
        sup_id = session.get('supervisor_id', -1)
        if existing:
            row_num = _find_row_by_id(ws_att, existing.get('id', 0))
            if row_num:
                headers = _get_headers(ws_att)
                ws_att.cell(row=row_num, column=headers.index('status') + 1, value=status)
                ws_att.cell(row=row_num, column=headers.index('notes') + 1, value=notes)
                ws_att.cell(row=row_num, column=headers.index('supervisor_id') + 1, value=sup_id)
        else:
            new_id = _next_id(ws_att)
            ws_att.append([new_id, employee_id, sup_id, record_time, status, notes])
        wb.save(EXCEL_PATH)
        wb.close()
    return jsonify({'message': 'OK', 'status': status})


# ===== EMPLOYEE ENDPOINTS =====

@app.route('/api/employees', methods=['GET'])
@login_required
def list_employees():
    service_id = request.args.get('service_id')
    stage_id = request.args.get('stage_id')
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Employees']
        employees = _sheet_to_dicts(ws)
        wb.close()
    result = []
    for e in employees:
        e['id'] = int(e['id'])
        e['service_id'] = _safe_int(e.get('service_id'))
        e['stage_id'] = _safe_int(e.get('stage_id'))
        if service_id and e['service_id'] != _safe_int(service_id):
            continue
        if stage_id and e['stage_id'] != _safe_int(stage_id):
            continue
        result.append(e)
    result.sort(key=lambda x: x['name'])
    return jsonify(result)


@app.route('/api/employees', methods=['POST'])
@login_required
def create_employee():
    data = request.json
    nfc_uid = data.get('nfc_uid', '').strip().upper()
    name = data.get('name', '').strip()
    if not nfc_uid or not name:
        return jsonify({'error': 'NFC UID and name required'}), 400
    service_id = _safe_int(data.get('service_id'))
    stage_id = _safe_int(data.get('stage_id'))
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Employees']
        # Check uniqueness
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and str(row[1]).upper() == nfc_uid:
                wb.close()
                return jsonify({'error': 'NFC UID already registered'}), 409
        new_id = _next_id(ws)
        ws.append([new_id, nfc_uid, name, data.get('birthdate', ''),
                   data.get('address', ''), data.get('phone', ''),
                   data.get('email', ''), data.get('class_name', ''),
                   data.get('parent_phone', ''), data.get('confession_father', ''),
                   '', service_id, stage_id, datetime.now().isoformat()])
        wb.save(EXCEL_PATH)
        # Read back the employee
        employee = _row_to_dict(ws, ws.max_row)
        employee['id'] = int(employee['id'])
        wb.close()
    return jsonify({'message': 'OK', 'employee': employee}), 201


@app.route('/api/employees/<int:emp_id>', methods=['GET'])
@login_required
def get_employee(emp_id):
    with _excel_lock:
        wb = _load_wb()
        ws_emp = wb['Employees']
        row_num = _find_row_by_id(ws_emp, emp_id)
        if not row_num:
            wb.close()
            return jsonify({'error': 'Not found'}), 404
        employee = _row_to_dict(ws_emp, row_num)
        employee['id'] = int(employee['id'])
        employee['service_id'] = _safe_int(employee.get('service_id'))
        employee['stage_id'] = _safe_int(employee.get('stage_id'))
        # Get attendance records
        ws_att = wb['Attendance']
        all_att = _sheet_to_dicts(ws_att)
        ws_sup = wb['Supervisors']
        all_sup = _sheet_to_dicts(ws_sup)
        # Get visits
        visits = []
        if 'Visits' in wb.sheetnames:
            visits = _sheet_to_dicts(wb['Visits'])
        wb.close()
    sup_map = {int(s['id']): s['name'] for s in all_sup}
    attendance = []
    for a in all_att:
        if int(a['employee_id']) == emp_id:
            rec = dict(a)
            rec['id'] = int(rec['id'])
            rec['employee_id'] = int(rec['employee_id'])
            rec['supervisor_id'] = _safe_int(rec['supervisor_id'])
            rec['supervisor_name'] = sup_map.get(_safe_int(a['supervisor_id']), '')
            attendance.append(rec)
    attendance.sort(key=lambda x: str(x['scan_time']), reverse=True)
    total_records = len(attendance)
    present_count = sum(1 for a in attendance if a['status'] == 'present')
    absent_count = sum(1 for a in attendance if a['status'] == 'absent')
    # Weekly analytics: last 12 Fridays
    weekly = []
    today = date.today()
    for i in range(12):
        friday = get_nearest_friday(today - timedelta(weeks=i))
        fri_start = str(friday) + ' 00:00:00'
        fri_end = str(friday) + ' 23:59:59'
        record = {}
        for a in attendance:
            scan_t = str(a.get('scan_time', ''))
            if fri_start <= scan_t and scan_t <= fri_end:
                record = a
                break
        weekly.append({
            'date': str(friday),
            'status': record.get('status', 'absent') if record else 'absent'
        })

    # Compute next birthday and last visit
    next_birthday = _get_next_birthday(employee.get('birthdate', ''))
    last_visit = _get_last_visit_date(emp_id, visits)

    return jsonify({
        'employee': employee,
        'attendance': attendance,
        'stats': {
            'total_records': total_records, 'present': present_count, 'absent': absent_count,
            'rate': round(float((present_count * 100.0 / total_records) if total_records > 0 else 0.0), 1) # pyre-ignore
        },
        'weekly': weekly,
        'next_birthday': next_birthday,
        'last_visit': last_visit,
    })


@app.route('/api/employees/<int:emp_id>', methods=['PUT'])
@login_required
def update_employee(emp_id):
    data = request.json
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Employees']
        row_num = _find_row_by_id(ws, emp_id)
        if not row_num:
            wb.close()
            return jsonify({'error': 'Not found'}), 404
        headers = _get_headers(ws)
        old = _row_to_dict(ws, row_num)
        field_map = {
            'name': data.get('name', old['name']),
            'birthdate': data.get('birthdate', old['birthdate']),
            'address': data.get('address', old['address']),
            'phone': data.get('phone', old['phone']),
            'parent_phone': data.get('parent_phone', old['parent_phone']),
            'confession_father': data.get('confession_father', old['confession_father']),
            'department': data.get('class_name', old['department']),
            'service_id': _safe_int(data.get('service_id', old.get('service_id'))),
            'stage_id': _safe_int(data.get('stage_id', old.get('stage_id'))),
        }
        for field, value in field_map.items():
            if field in headers:
                ws.cell(row=row_num, column=headers.index(field) + 1, value=value)
        wb.save(EXCEL_PATH)
        updated = _row_to_dict(ws, row_num)
        updated['id'] = int(updated['id'])
        wb.close()
    return jsonify({'message': 'OK', 'employee': updated})


@app.route('/api/employees/<int:emp_id>', methods=['DELETE'])
@login_required
def delete_employee(emp_id):
    with _excel_lock:
        wb = _load_wb()
        # Delete attendance records
        ws_att = wb['Attendance']
        rows_to_delete = []
        for row_num in range(2, ws_att.max_row + 1):
            val = ws_att.cell(row=row_num, column=2).value
            if val is not None and int(val) == emp_id:
                rows_to_delete.append(row_num)
        for row_num in reversed(rows_to_delete):
            ws_att.delete_rows(row_num)
        # Delete visits
        if 'Visits' in wb.sheetnames:
            ws_vis = wb['Visits']
            vis_headers = _get_headers(ws_vis)
            emp_col = vis_headers.index('employee_id') + 1
            rows_to_delete = []
            for row_num in range(2, ws_vis.max_row + 1):
                val = ws_vis.cell(row=row_num, column=emp_col).value
                if val is not None and _safe_int(val) == emp_id:
                    rows_to_delete.append(row_num)
            for row_num in reversed(rows_to_delete):
                ws_vis.delete_rows(row_num)
        # Delete employee
        ws_emp = wb['Employees']
        emp_row = _find_row_by_id(ws_emp, emp_id)
        if emp_row:
            ws_emp.delete_rows(emp_row)
        wb.save(EXCEL_PATH)
        wb.close()
    return jsonify({'message': 'OK'})


# ===== VISITS (افتقادات) ENDPOINTS =====

@app.route('/api/visits', methods=['GET'])
@login_required
def list_visits():
    """List visits, optionally filtered by employee_id, service_id, stage_id."""
    employee_id = request.args.get('employee_id')
    service_id = request.args.get('service_id')
    stage_id = request.args.get('stage_id')
    with _excel_lock:
        wb = _load_wb()
        visits = _sheet_to_dicts(wb['Visits']) if 'Visits' in wb.sheetnames else []
        employees = _sheet_to_dicts(wb['Employees'])
        supervisors = _sheet_to_dicts(wb['Supervisors'])
        wb.close()
    emp_map = {_safe_int(e['id']): e for e in employees}
    sup_map = {_safe_int(s['id']): s['name'] for s in supervisors}

    result = []
    for v in visits:
        eid = _safe_int(v.get('employee_id'))
        if employee_id and eid != _safe_int(employee_id):
            continue
        emp = emp_map.get(eid, {})
        if service_id and _safe_int(emp.get('service_id')) != _safe_int(service_id):
            continue
        if stage_id and _safe_int(emp.get('stage_id')) != _safe_int(stage_id):
            continue
        result.append({
            'id': _safe_int(v['id']),
            'employee_id': eid,
            'employee_name': emp.get('name', ''),
            'supervisor_id': _safe_int(v.get('supervisor_id')),
            'supervisor_name': sup_map.get(_safe_int(v.get('supervisor_id')), ''),
            'visit_date': v.get('visit_date', ''),
            'notes': v.get('notes', ''),
            'created_at': v.get('created_at', ''),
        })
    result.sort(key=lambda x: str(x.get('visit_date', '')), reverse=True)
    return jsonify(result)


@app.route('/api/visits', methods=['POST'])
@login_required
def create_visit():
    data = request.json
    employee_id = _safe_int(data.get('employee_id'))
    visit_date = data.get('visit_date', date.today().isoformat())
    notes = data.get('notes', '')
    if not employee_id:
        return jsonify({'error': 'Employee ID required'}), 400
    sup_id = session.get('supervisor_id', -1)
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Visits']
        new_id = _next_id(ws)
        ws.append([new_id, employee_id, sup_id, visit_date, notes, datetime.now().isoformat()])
        wb.save(EXCEL_PATH)
        wb.close()
    return jsonify({
        'message': 'OK',
        'visit': {
            'id': new_id, 'employee_id': employee_id,
            'supervisor_id': sup_id, 'visit_date': visit_date, 'notes': notes
        }
    }), 201


@app.route('/api/visits/<int:visit_id>', methods=['DELETE'])
@login_required
def delete_visit(visit_id):
    with _excel_lock:
        wb = _load_wb()
        ws = wb['Visits']
        row = _find_row_by_id(ws, visit_id)
        if row:
            ws.delete_rows(row)
        wb.save(EXCEL_PATH)
        wb.close()
    return jsonify({'message': 'OK'})


# ===== BIRTHDAYS ENDPOINT =====

@app.route('/api/birthdays', methods=['GET'])
@login_required
def upcoming_birthdays():
    """Return employees with birthdays in the next N days (default 30)."""
    days_ahead = int(request.args.get('days', 30))
    service_id = request.args.get('service_id')
    stage_id = request.args.get('stage_id')
    with _excel_lock:
        wb = _load_wb()
        employees = _sheet_to_dicts(wb['Employees'])
        wb.close()
    today = date.today()
    result = []
    for emp in employees:
        if service_id and _safe_int(emp.get('service_id')) != _safe_int(service_id):
            continue
        if stage_id and _safe_int(emp.get('stage_id')) != _safe_int(stage_id):
            continue
        bd_str = str(emp.get('birthdate', '')).strip()
        if not bd_str:
            continue
        try:
            bd_str_clean = bd_str.split('T')[0].split(' ')[0]
            bd = datetime.strptime(bd_str_clean, '%Y-%m-%d').date()
        except (ValueError, TypeError):
            continue
        this_year_bd = bd.replace(year=today.year)
        if this_year_bd < today:
            this_year_bd = bd.replace(year=today.year + 1)
        days_until = (this_year_bd - today).days
        if days_until <= days_ahead:
            age = this_year_bd.year - bd.year
            result.append({
                'id': _safe_int(emp['id']),
                'name': emp['name'],
                'birthdate': bd_str,
                'next_birthday': this_year_bd.isoformat(),
                'days_until': days_until,
                'age': age,
                'phone': emp.get('phone', ''),
                'department': emp.get('department', ''),
                'service_id': _safe_int(emp.get('service_id')),
                'stage_id': _safe_int(emp.get('stage_id')),
            })
    result.sort(key=lambda x: x['days_until'])
    return jsonify(result)


# ===== DASHBOARD =====

@app.route('/api/dashboard', methods=['GET'])
@login_required
def dashboard():
    target_date = request.args.get('date', date.today().isoformat())
    service_id = request.args.get('service_id')
    stage_id = request.args.get('stage_id')
    with _excel_lock:
        wb = _load_wb()
        employees = _sheet_to_dicts(wb['Employees'])
        attendance = _sheet_to_dicts(wb['Attendance'])
        supervisors = _sheet_to_dicts(wb['Supervisors'])
        wb.close()
    # Filter employees by service/stage
    if service_id:
        employees = [e for e in employees if _safe_int(e.get('service_id')) == _safe_int(service_id)]
    if stage_id:
        employees = [e for e in employees if _safe_int(e.get('stage_id')) == _safe_int(stage_id)]
    emp_ids = {int(e['id']) for e in employees}
    total_employees = len(employees)
    total_supervisors = len(supervisors)
    day_start = target_date + ' 00:00:00'
    day_end = target_date + ' 23:59:59'
    day_records = [a for a in attendance
                   if day_start <= str(a['scan_time']) <= day_end
                   and int(a['employee_id']) in emp_ids]
    today_present = sum(1 for a in day_records if a['status'] == 'present')
    today_absent = sum(1 for a in day_records if a['status'] == 'absent')
    emp_map = {int(e['id']): e['name'] for e in employees}
    sup_map = {int(s['id']): s['name'] for s in supervisors}
    recent = []
    sorted_records = sorted(day_records, key=lambda x: str(x.get('scan_time', '')), reverse=True)
    for i in range(min(20, len(sorted_records))):
        a = sorted_records[i]
        recent.append({
            'employee_id': int(a.get('employee_id', 0)),
            'employee_name': emp_map.get(int(a.get('employee_id', 0)), ''),
            'nfc_uid': next((e.get('nfc_uid', '') for e in employees if int(e['id']) == int(a['employee_id'])), ''),
            'supervisor_name': sup_map.get(_safe_int(a['supervisor_id']), ''),
            'scan_time': a['scan_time'], 'status': a['status']
        })
    return jsonify({
        'total_employees': total_employees, 'today_present': today_present,
        'today_absent': today_absent,
        'today_not_scanned': max(0, total_employees - today_present - today_absent),
        'total_supervisors': total_supervisors,
        'recent_scans': recent, 'date': target_date
    })


# ===== ATTENDANCE BY DATE =====

@app.route('/api/attendance/date', methods=['GET'])
@login_required
def attendance_by_date():
    target_date = request.args.get('date', date.today().isoformat())
    service_id = request.args.get('service_id')
    stage_id = request.args.get('stage_id')
    with _excel_lock:
        wb = _load_wb()
        employees = _sheet_to_dicts(wb['Employees'])
        attendance = _sheet_to_dicts(wb['Attendance'])
        supervisors = _sheet_to_dicts(wb['Supervisors'])
        wb.close()
    if service_id:
        employees = [e for e in employees if _safe_int(e.get('service_id')) == _safe_int(service_id)]
    if stage_id:
        employees = [e for e in employees if _safe_int(e.get('stage_id')) == _safe_int(stage_id)]
    day_start = target_date + ' 00:00:00'
    day_end = target_date + ' 23:59:59'
    sup_map = {int(s['id']): s['name'] for s in supervisors}
    result = []
    for emp in sorted(employees, key=lambda x: x.get('name', '')):
        record = {}
        for a in attendance:
            scan_t = str(a.get('scan_time', ''))
            if (int(a.get('employee_id', 0)) == int(emp.get('id', 0)) and
                    day_start <= scan_t and scan_t <= day_end):
                record = a
                break
        emp_dict = dict(emp)
        emp_dict['id'] = int(emp_dict.get('id', 0))
        result.append({
            'employee': emp_dict,
            'status': record.get('status', 'not_scanned') if record else 'not_scanned',
            'scan_time': record.get('scan_time') if record else None,
            'supervisor': sup_map.get(_safe_int(record.get('supervisor_id', 0)), '') if record else None
        })
    return jsonify(result)


@app.route('/api/attendance/today', methods=['GET'])
@login_required
def today_attendance():
    with _excel_lock:
        wb = _load_wb()
        employees = _sheet_to_dicts(wb['Employees'])
        attendance = _sheet_to_dicts(wb['Attendance'])
        supervisors = _sheet_to_dicts(wb['Supervisors'])
        wb.close()
    today_str = date.today().isoformat()
    day_start = today_str + ' 00:00:00'
    day_end = today_str + ' 23:59:59'
    sup_map = {int(s['id']): s['name'] for s in supervisors}
    result = []
    for emp in sorted(employees, key=lambda x: x.get('name', '')):
        record = {}
        for a in attendance:
            scan_t = str(a.get('scan_time', ''))
            if (int(a.get('employee_id', 0)) == int(emp.get('id', 0)) and
                    day_start <= scan_t and scan_t <= day_end):
                record = a
                break
        emp_dict = dict(emp)
        emp_dict['id'] = int(emp_dict.get('id', 0))
        result.append({
            'employee': emp_dict,
            'status': record.get('status', 'not_scanned') if record else 'not_scanned',
            'scan_time': record.get('scan_time') if record else None,
            'supervisor': sup_map.get(_safe_int(record.get('supervisor_id', 0)), '') if record else None
        })
    return jsonify(result)


# ===== ANALYTICS =====

@app.route('/api/analytics', methods=['GET'])
@login_required
def analytics():
    weeks = int(request.args.get('weeks', 12))
    service_id = request.args.get('service_id')
    stage_id = request.args.get('stage_id')
    with _excel_lock:
        wb = _load_wb()
        employees = _sheet_to_dicts(wb['Employees'])
        attendance = _sheet_to_dicts(wb['Attendance'])
        wb.close()
    if service_id:
        employees = [e for e in employees if _safe_int(e.get('service_id')) == _safe_int(service_id)]
    if stage_id:
        employees = [e for e in employees if _safe_int(e.get('stage_id')) == _safe_int(stage_id)]
    emp_ids = {int(e['id']) for e in employees}
    attendance = [a for a in attendance if int(a['employee_id']) in emp_ids]
    total_employees = len(employees)
    # Weekly attendance trend
    weekly_trend = []
    today = date.today()
    for i in range(weeks):
        friday = get_nearest_friday(today - timedelta(weeks=i))
        fri_start = friday.isoformat() + ' 00:00:00'
        fri_end = friday.isoformat() + ' 23:59:59'
        day_records = [a for a in attendance if fri_start <= str(a.get('scan_time', '')) <= fri_end]
        present = sum(1 for a in day_records if a.get('status') == 'present')
        absent = sum(1 for a in day_records if a.get('status') == 'absent')
        weekly_trend.append({
            'date': friday.isoformat(), 'present': present, 'absent': absent, 'total': total_employees
        })
    # Per-attendee stats
    attendee_stats = []
    for emp in employees:
        emp_att = [a for a in attendance if int(a.get('employee_id', 0)) == int(emp.get('id', 0))]
        present = sum(1 for a in emp_att if a.get('status') == 'present')
        total = len(emp_att)
        rate = round(float((present * 100.0 / total) if total > 0 else 0.0), 1) # pyre-ignore
        attendee_stats.append({
            'id': int(emp['id']), 'name': emp['name'], 'class_name': emp.get('department', ''),
            'present': present, 'total': total, 'rate': rate
        })
    attendee_stats.sort(key=lambda x: x['rate'], reverse=True)
    # Class breakdown
    classes = {}
    for emp in employees:
        cls = emp.get('department', '') or 'Other'
        if cls not in classes:
            classes[cls] = {'name': cls, 'count': 0, 'total_present': 0, 'total_records': 0}
        classes[cls]['count'] += 1
        emp_att = [a for a in attendance if int(a.get('employee_id', 0)) == int(emp.get('id', 0))]
        present = sum(1 for a in emp_att if a.get('status') == 'present')
        classes[cls]['total_present'] += present
        classes[cls]['total_records'] += len(emp_att)
    class_stats = []
    for cls_data in classes.values():
        rate = round(float((cls_data['total_present'] * 100.0 / cls_data['total_records']) if cls_data['total_records'] > 0 else 0.0), 1) # pyre-ignore
        class_stats.append({**cls_data, 'rate': rate})
    # Overall
    total_attendance = len(attendance)
    total_present = sum(1 for a in attendance if a.get('status') == 'present')
    overall_rate = round(float((total_present * 100.0 / total_attendance) if total_attendance > 0 else 0.0), 1) # pyre-ignore
    return jsonify({
        'total_employees': total_employees, 'overall_rate': overall_rate,
        'total_attendance_records': total_attendance, 'total_present': total_present,
        'weekly_trend': weekly_trend, 'attendee_stats': attendee_stats, 'class_stats': class_stats
    })


@app.route('/api/attendance/report', methods=['GET'])
@login_required
def attendance_report():
    start_date = request.args.get('start', (date.today() - timedelta(days=90)).isoformat())
    end_date = request.args.get('end', date.today().isoformat())
    with _excel_lock:
        wb = _load_wb()
        attendance = _sheet_to_dicts(wb['Attendance'])
        employees = _sheet_to_dicts(wb['Employees'])
        supervisors = _sheet_to_dicts(wb['Supervisors'])
        wb.close()
    emp_map = {int(e['id']): e for e in employees}
    sup_map = {int(s['id']): s['name'] for s in supervisors}
    records = []
    for a in attendance:
        scan_date_str = str(a['scan_time']).split(' ')[0] if a['scan_time'] else ''
        if start_date <= scan_date_str <= end_date:
            emp = emp_map.get(int(a['employee_id']), {})
            records.append({
                'id': int(a['id']), 'employee_id': int(a['employee_id']),
                'supervisor_id': _safe_int(a['supervisor_id']),
                'scan_time': a['scan_time'], 'status': a['status'],
                'notes': a.get('notes', ''),
                'employee_name': emp.get('name', ''),
                'nfc_uid': emp.get('nfc_uid', ''),
                'supervisor_name': sup_map.get(_safe_int(a['supervisor_id']), '')
            })
    records.sort(key=lambda x: str(x['scan_time']), reverse=True)
    return jsonify(records)


# ===== MAIN =====

if __name__ == '__main__':
    import socket
    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    try:
        s.connect(('8.8.8.8', 80))
        local_ip = s.getsockname()[0]
    except Exception:
        local_ip = '0.0.0.0'
    finally:
        s.close()

    cert_file = os.path.join(RUN_DIR, 'cert.pem')
    key_file = os.path.join(RUN_DIR, 'key.pem')
    if not os.path.exists(cert_file) or not os.path.exists(key_file):
        from OpenSSL import crypto # pyre-ignore
        k = crypto.PKey()
        k.generate_key(crypto.TYPE_RSA, 2048)
        cert = crypto.X509()
        cert.get_subject().CN = local_ip
        cert.set_serial_number(1000)
        cert.gmtime_adj_notBefore(0)
        cert.gmtime_adj_notAfter(365 * 24 * 60 * 60)
        cert.set_issuer(cert.get_subject())
        cert.set_pubkey(k)
        cert.sign(k, 'sha256')
        with open(cert_file, 'wb') as f:
            f.write(crypto.dump_certificate(crypto.FILETYPE_PEM, cert))
        with open(key_file, 'wb') as f:
            f.write(crypto.dump_privatekey(crypto.FILETYPE_PEM, k))

    print("=" * 50)
    print("  NFC Attendance System - Sunday School")
    print("=" * 50)
    print(f"  Phone:   https://{local_ip}:5000")
    print("=" * 50)

    # Start Flask in a background thread
    flask_thread = threading.Thread(
        target=lambda: app.run(
            host='0.0.0.0', port=5000, debug=False,
            ssl_context=(cert_file, key_file),
            use_reloader=False
        ),
        daemon=True
    )
    flask_thread.start()

    # Auto-login as master admin for the pywebview desktop session
    # We do this by injecting the admin secret via a custom JS bridge
    import webview # pyre-ignore

    class MasterAdminApi:
        def get_admin_secret(self):
            return MASTER_ADMIN_SECRET

    api = MasterAdminApi()
    webview.create_window(
        'مدارس الاحد - NFC Attendance',
        'https://localhost:5000',
        width=420, height=780,
        resizable=True,
        min_size=(360, 600),
        js_api=api
    )
    webview.start(ssl=True)
